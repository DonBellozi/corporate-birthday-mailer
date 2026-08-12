import re
from datetime import date, datetime
from pathlib import Path
from typing import Any
from openpyxl import load_workbook

MONTHS_RU = {
    "январь":1,"февраль":2,"март":3,"апрель":4,"май":5,"июнь":6,
    "июль":7,"август":8,"сентябрь":9,"октябрь":10,"ноябрь":11,"декабрь":12,
}

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()

def normalize_key(value: Any) -> str:
    return normalize_text(value).lower().replace("ё", "е")

def parse_birthday(value: Any):
    if value is None or not normalize_text(value):
        return None
    if isinstance(value, (date, datetime)):
        return value.day, value.month

    text = normalize_text(value)
    if "," in text:
        day_text, month_text = text.split(",", 1)
        day = int(day_text.strip())
        month = MONTHS_RU[normalize_key(month_text)]
        date(2000, month, day)
        return day, month

    for fmt in ("%d.%m", "%d.%m.%Y", "%d/%m", "%d/%m/%Y"):
        try:
            d = datetime.strptime(text, fmt)
            return d.day, d.month
        except ValueError:
            pass
    raise ValueError(f"Неизвестный формат даты рождения: {value!r}")

def detect_gender(fio: str, explicit: str = "") -> str:
    x = normalize_key(explicit)
    if x in {"м", "муж", "мужской", "male"}:
        return "male"
    if x in {"ж", "жен", "женский", "female"}:
        return "female"

    parts = normalize_text(fio).split()
    patronymic = parts[2].lower() if len(parts) >= 3 else ""
    if patronymic.endswith(("ович", "евич", "ич")):
        return "male"
    if patronymic.endswith(("овна", "евна", "ична", "инична")):
        return "female"
    return "unknown"

def build_header_map(ws, top_row: int, second_row: int | None):
    result = {}
    for col in range(1, ws.max_column + 1):
        top = normalize_text(ws.cell(top_row, col).value)
        bottom = normalize_text(ws.cell(second_row, col).value) if second_row else ""
        for value in [top, bottom, f"{top} {bottom}".strip()]:
            if value:
                result[normalize_key(value)] = col
    return result


HEADER_FIO_ALIASES = {
    normalize_key("Сотрудник.Физическое лицо.ФИО"),
    normalize_key("Физическое лицо.ФИО"),
    normalize_key("ФИО"),
}

HEADER_HIDE_ALIASES = {
    normalize_key("Сотрудник.Скрыть день рождения (Сотрудники)"),
}

HEADER_POSITION_ALIASES = {
    normalize_key("Должность"),
}


def detect_header_rows(ws, configured_top=2, configured_second=3):
    """
    Отчет 1С использует две строки заголовков:
      строка 2 — основные заголовки;
      строка 3 — уточняющие подзаголовки;
      строка 4 — начало содержимого.

    Сначала проверяем настроенные строки. Если они были случайно изменены
    в веб-интерфейсе или остались от старой версии, автоматически ищем
    строку с заголовком ФИО среди первых 15 строк.
    """
    candidates = []

    try:
        top = int(configured_top)
        second = int(configured_second) if configured_second else top + 1
        candidates.append((top, second))
    except Exception:
        pass

    # Фактическая структура текущего отчета 1С.
    candidates.append((2, 3))

    # Дополнительный автопоиск.
    for row in range(1, min(ws.max_row, 15) + 1):
        values = {
            normalize_key(ws.cell(row, col).value)
            for col in range(1, ws.max_column + 1)
            if normalize_text(ws.cell(row, col).value)
        }

        if values & HEADER_FIO_ALIASES:
            candidates.append((row, row + 1 if row < ws.max_row else None))

    seen = set()
    for top, second in candidates:
        if not top or top < 1 or top > ws.max_row:
            continue
        key = (top, second)
        if key in seen:
            continue
        seen.add(key)

        headers = build_header_map(ws, top, second)
        header_names = set(headers.keys())

        # Заголовок ФИО обязателен. Дополнительные признаки защищают
        # от случайного совпадения текста "ФИО" внутри содержимого.
        has_fio = bool(header_names & HEADER_FIO_ALIASES)
        has_position = bool(header_names & HEADER_POSITION_ALIASES)
        has_hide = bool(header_names & HEADER_HIDE_ALIASES)

        if has_fio and (has_position or has_hide):
            return top, second

    raise ValueError(
        "Не удалось определить строки заголовков XLSX. "
        "Ожидался заголовок «Сотрудник.Физическое лицо.ФИО» "
        "во 2-й строке и подзаголовки в 3-й."
    )

def update_department_stack(stack: list[str], level: int, name: str) -> list[str]:
    """
    Обновляет текущий путь подразделений по Excel outlineLevel.

    В выгрузке 1С строки подразделений находятся в колонке A и объединены
    по ширине отчета. Уровень группировки строки задает место подразделения
    в иерархии.
    """
    name = normalize_text(name)
    if not name:
        return stack

    level = max(0, int(level or 0))

    # Оставляем только родителей текущего уровня.
    stack = stack[:level]

    # В отчете встречаются технические повторы одного подразделения
    # на соседних уровнях. Не добавляем одинаковое имя второй раз подряд.
    if stack and normalize_key(stack[-1]) == normalize_key(name):
        return stack

    stack.append(name)
    return stack


def build_position_with_departments(department_stack: list[str], position: str) -> str:
    """
    Формирует полное исходное представление для справочника должностей:
      Центральный Аппарат / Департамент ... / Отдел ...  Должность
    """
    parts = []
    for item in department_stack:
        item = normalize_text(item)
        if not item:
            continue
        # На всякий случай дополнительно убираем последовательные дубли.
        if parts and normalize_key(parts[-1]) == normalize_key(item):
            continue
        parts.append(item)

    path = " / ".join(parts)
    position = normalize_text(position)

    if path and position:
        return f"{path}  {position}"
    return path or position


def find_col(headers, expected: str, required=True, aliases=None):
    aliases = aliases or []
    names = [expected, *aliases]

    # Сначала только точные совпадения. Это важно для заголовков вида
    # "Сотрудник....", которых в отчете 1С несколько.
    for name in names:
        target = normalize_key(name)
        if target and target in headers:
            return headers[target]

    # Затем мягкий поиск, но только если получился ровно один кандидат.
    for name in names:
        target = normalize_key(name)
        if not target:
            continue
        candidates = [(header, col) for header, col in headers.items()
                      if target in header or header in target]
        unique_cols = sorted({col for _, col in candidates})
        if len(unique_cols) == 1:
            return unique_cols[0]

    if required:
        raise ValueError(
            f"Не найдена колонка: {expected}. "
            f"Доступные заголовки: {', '.join(sorted(headers.keys()))}"
        )
    return None

def parse_workbook(path: str | Path, cfg: dict[str, str]):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Не доверяем слепо сохраненным номерам строк в настройках.
    # Для текущей выгрузки 1С это 2-я и 3-я строки, но при необходимости
    # парсер найдет их сам.
    h1, h2 = detect_header_rows(
        ws,
        cfg.get("xlsx_header_row", "2"),
        cfg.get("xlsx_second_header_row", "3"),
    )

    # Содержимое начинается сразу после второй строки заголовков.
    # Для текущего отчета: 2–3 заголовки, данные с 4-й.
    data_row = (h2 + 1) if h2 else (h1 + 1)
    headers = build_header_map(ws, h1, h2)

    fio_col = find_col(
        headers,
        cfg["xlsx_fio_column"],
        aliases=["Сотрудник.Физическое лицо.ФИО", "Физическое лицо.ФИО", "ФИО"],
    )
    birthday_col = find_col(
        headers,
        cfg["xlsx_birthday_column"],
        aliases=[
            "Дата рождения.День, Дата рождения.Название месяца",
            "День месяц года",
        ],
    )
    position_col = find_col(
        headers,
        cfg.get("xlsx_position_column", ""),
        False,
        aliases=["Должность"],
    )
    hide_col = find_col(
        headers,
        cfg.get("xlsx_hide_column", ""),
        False,
        aliases=["Сотрудник.Скрыть день рождения (Сотрудники)"],
    )
    id_col = find_col(
        headers,
        cfg.get("xlsx_id_column", ""),
        False,
        aliases=["СНИЛС"],
    )
    gender_col = find_col(
        headers,
        cfg.get("xlsx_gender_column", ""),
        False,
        aliases=["Физическое лицо.Пол", "Пол"],
    )
    state_col = find_col(
        headers,
        cfg.get("xlsx_state_column", ""),
        False,
        aliases=["Состояние"],
    )

    result = []
    department_stack: list[str] = []

    for row in range(data_row, ws.max_row + 1):
        fio = normalize_text(ws.cell(row, fio_col).value)

        # Строка без ФИО, но со значением в первой колонке — это строка
        # подразделения. В ней колонка A объединена по ширине отчета.
        # outlineLevel определяет глубину подразделения.
        if not fio:
            department_name = normalize_text(ws.cell(row, 1).value)
            if department_name:
                department_stack = update_department_stack(
                    department_stack,
                    ws.row_dimensions[row].outlineLevel,
                    department_name,
                )
            continue

        raw_birthday = ws.cell(row, birthday_col).value
        try:
            birthday = parse_birthday(raw_birthday)
        except Exception as exc:
            result.append({"fio": fio, "error": str(exc)})
            continue
        if not birthday:
            continue

        position = normalize_text(ws.cell(row, position_col).value) if position_col else ""
        source_position = build_position_with_departments(department_stack, position)

        hidden = normalize_key(ws.cell(row, hide_col).value) == "да" if hide_col else False
        explicit_gender = normalize_text(ws.cell(row, gender_col).value) if gender_col else ""
        gender = detect_gender(fio, explicit_gender)
        employee_state = normalize_text(ws.cell(row, state_col).value) if state_col else ""
        employee_key = normalize_text(ws.cell(row, id_col).value) if id_col else ""
        if not employee_key:
            employee_key = normalize_key(f"{fio}|{birthday[0]:02d}.{birthday[1]:02d}")

        result.append({
            "employee_key": employee_key,
            "fio": fio,
            "birthday_day": birthday[0],
            "birthday_month": birthday[1],
            "gender": gender,
            "employee_state": employee_state,
            "hide_birthday": hidden,
            "source_position": source_position,
            "error": None,
        })
    return result
